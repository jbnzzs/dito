import ast
import os

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import pycountry


class Command(BaseCommand):
    help = "Importa imagens e descrições a partir de um arquivo Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument(
            "arquivo",
            type=str,
            help="Caminho para o arquivo Excel (.xlsx)",
        )
        parser.add_argument(
            "--atualizar",
            action="store_true",
            help="Atualizar registros existentes (padrão: pular duplicados)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl não está instalado. Execute: pip install openpyxl"
            )

        from core.models import Descricao, Imagem, StatusWorkflow, Trecho, Usuario

        arquivo = options["arquivo"]
        atualizar = options["atualizar"]

        if not os.path.exists(arquivo):
            raise CommandError(f"Arquivo não encontrado: {arquivo}")

        # Status inicial para todas as imagens importadas
        try:
            status_inicial = StatusWorkflow.objects.get(slug="liberado-descricao")
        except StatusWorkflow.DoesNotExist:
            raise CommandError(
                "Status 'Liberado para descrição' não encontrado. "
                "Execute: python manage.py seed_status"
            )

        self.stdout.write(f"Carregando: {arquivo}")
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active

        # Ler cabeçalhos da primeira linha
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.stdout.write(f"Colunas: {headers}\n")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        self.stdout.write(f"Total de registros no arquivo: {total}\n")

        criadas = atualizadas = puladas = erros = 0

        for i, row in enumerate(rows, 1):
            data = dict(zip(headers, row))
            retranca = data.get("retranca")

            if not retranca:
                self.stdout.write(self.style.WARNING(f"  Linha {i+1}: retranca vazia — pulando"))
                puladas += 1
                continue

            try:
                with transaction.atomic():

                    # Verificar duplicado
                    imagem_existente = Imagem.objects.filter(retranca=retranca).first()
                    if imagem_existente and not atualizar:
                        puladas += 1
                        continue

                    # Buscar usuário responsável pelo username
                    responsavel = None
                    username = data.get("usuario")
                    if username:
                        try:
                            responsavel = Usuario.objects.get(username=username)
                        except Usuario.DoesNotExist:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"  Linha {i+1}: usuário '{username}' não encontrado"
                                )
                            )

                    # Processar etapa ("Etapa: AD" → "AD")
                    etapa_raw = str(data.get("etapa") or "AD")
                    etapa = etapa_raw.replace("Etapa:", "").strip()
                    etapas_validas = [e[0] for e in Imagem.Etapa.choices]
                    if etapa not in etapas_validas:
                        etapa = Imagem.Etapa.AD

                    # Processar caminho e nome do arquivo
                    img_file = str(data.get("img_file") or "")
                    nome_arquivo = os.path.basename(img_file) if img_file else ""

                    campos = {
                        "nome_obra": str(data.get("colecao") or ""),
                        "componente_curricular": str(data.get("disciplina") or ""),
                        "volume_ano_modulo": str(data.get("volume") or ""),
                        "capitulo_unidade": str(data.get("capitulo") or ""),
                        "etapa": etapa,
                        "nome_arquivo": nome_arquivo,
                        "caminho_arquivo": img_file,
                        "status": status_inicial,
                        "responsavel": responsavel,
                        "ativo": True,
                    }

                    if imagem_existente and atualizar:
                        for campo, valor in campos.items():
                            setattr(imagem_existente, campo, valor)
                        imagem_existente.save()
                        imagem = imagem_existente
                        atualizadas += 1
                        self.stdout.write(f"  ~ Atualizada: {retranca}")
                    else:
                        imagem = Imagem.objects.create(retranca=retranca, **campos)
                        criadas += 1
                        self.stdout.write(self.style.SUCCESS(f"  ✓ Criada: {retranca}"))

                    # Processar descrição (JSON com trechos e idiomas)
                    descricao_raw = data.get("descricao")
                    if descricao_raw:
                        try:
                            trechos_data = ast.literal_eval(str(descricao_raw))

                            if isinstance(trechos_data, list) and trechos_data:
                                descricao, _ = Descricao.objects.get_or_create(
                                    imagem=imagem,
                                    defaults={
                                        "descritor": responsavel,
                                        "descritor_bloqueado": False,
                                    },
                                )

                                # Limpar trechos antigos se estiver atualizando
                                if atualizar:
                                    descricao.trechos.all().delete()

                                for ordem, trecho_data in enumerate(trechos_data, 1):
                                    lang_code = trecho_data.get("lang", "pt-BR")
                                    texto = trecho_data.get("text", "")

                                    # Mapear código de idioma via pycountry
                                    alpha_2 = lang_code.split("-")[0]
                                    try:
                                        lang = pycountry.languages.get(alpha_2=alpha_2)
                                        idioma_codigo = (
                                            lang.alpha_3
                                            if lang and hasattr(lang, "alpha_3")
                                            else "und"
                                        )
                                        idioma_nome = lang.name if lang else lang_code
                                    except Exception:
                                        idioma_codigo = "und"
                                        idioma_nome = lang_code

                                    Trecho.objects.create(
                                        descricao=descricao,
                                        ordem=ordem,
                                        texto=texto,
                                        idioma_codigo=idioma_codigo,
                                        idioma_nome=idioma_nome,
                                    )

                        except (ValueError, SyntaxError) as e:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"  Linha {i+1}: erro ao processar descrição — {e}"
                                )
                            )

            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f"  Linha {i+1} ({retranca}): erro — {e}")
                )
                erros += 1

        wb.close()

        self.stdout.write(
            self.style.SUCCESS(
                f"\n{'='*40}\n"
                f"Importação concluída:\n"
                f"  ✓ Criadas:    {criadas}\n"
                f"  ~ Atualizadas: {atualizadas}\n"
                f"  - Puladas:    {puladas}\n"
                f"  ✗ Erros:      {erros}\n"
                f"{'='*40}"
            )
        )